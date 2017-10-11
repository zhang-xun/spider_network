#!/usr/bin/env python
# -*- coding: utf-8 -*-

from openpyxl import Workbook
from openpyxl import load_workbook
class WriterExcel(Workbook):
    def __init__(self, excel_savename,sheetname):
        self.excel_savename = excel_savename
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws1 = self.wb.create_sheet(sheetname)

    def append_title(self,title):
        self.ws1.append(title) 
        print("success append title")
    
    def append(self,list):
        for i in list:
            self.ws1.append(i.split("\t"))
        print("success append content")

    def save(self):
        self.wb.save(self.excel_savename)

    
class ReaderExcel(Workbook):
    def __init__(self,excel_readname,sheetname):
        self.wb = load_workbook(excel_readname)
        self.ws = self.wb[sheetname]
        


    def read(self):
        for row in self.ws.rows:
            for cell in row:
                print(cell.value,end="")
            print()

        print(tuple(self.ws.columns))

    def count_form_FL_to(self):
        exmpt_number = 0 
        for index,cell in enumerate(tuple(self.ws.columns)[0]):
            #print(len(cell.value))
            #print(cell.value[:-1]+":after")
            #print(int(tuple(self.ws.columns)[7][3].value))
            #print(self.ws['C3'].value)
            if cell.value ==  self.ws['C3'].value :
                
                exmpt_number += int(tuple(self.ws.columns)[7][index].value)
        print("FLORIDA\t",exmpt_number)

    def sum_of_migration(self):
        migration_number = 0
        for cells in tuple(self.ws.rows)[1:]:
            migration_number += int(cells[7].value)
            migration_number += int(cells[8].value)
        print("Total\t\t",migration_number)







def main():
    print("before start")
    writer = WriterExcel("us_state.xlsx","USA_state_condition")
    print("work __init__")
    with open("shuai.txt") as f:
        writer.append_title(f.readline().split("\t")) 
        writer.append(f.readlines())
    print("open txt file done")
    writer.save()
    
def reader_main():
    print("STATE"+"\t"*2+"TOTAL")
    reader = ReaderExcel("us_state.xlsx","USA_state_condition")
    #reader.read()
    reader.count_form_FL_to()
    reader.sum_of_migration()



if __name__ == "__main__":
    #main()
    reader_main()
    
